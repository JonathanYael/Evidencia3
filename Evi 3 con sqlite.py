import openpyxl
import csv
import os
import sqlite3
from sqlite3 import Error
import sys
diccionario_ejemplar= dict()
diccionario_ejemplar_autor= dict()
diccionario_ejemplar_genero= dict()
diccionario_ejemplar_año= dict()
encabezados = ["id", "titulo", "nombres de autor", "apellidos de autor", "genero", "año publicacion", "isbn", "fecha adquisicion"]


if os.path.exists("biblioteca x.db"):
    print("se encontró un archivo de base de datos")
else:
    print("no se encontró un archivo previo, se procede a crear uno...")
    try:
        with sqlite3.connect("biblioteca x.db") as conn:
            cursor = conn.cursor()
            cursor.execute("CREATE TABLE autor (id INTEGER PRIMARY KEY autoincrement, nombres TEXT NOT NULL, apellidos TEXT NOT NULL);")
            cursor.execute("CREATE TABLE genero (id INTEGER PRIMARY KEY autoincrement, nombre TEXT NOT NULL);")
            cursor.execute("CREATE TABLE ejemplar (id INTEGER PRIMARY KEY autoincrement, titulo TEXT NOT NULL, autor INTEGER NOT NULL, genero INTEGER NOT NULL, año_pub INTEGER NOT NULL, isbn INTEGER NOT NULL, fech_adq TIMESTAMP NOT NULL, FOREIGN KEY(autor) REFERENCES autor(id), FOREIGN KEY(genero) REFERENCES genero(id));")   
            print("\nse ha creado la base de datos")
    except Error as e:
        print (e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")  

while True:
    print("\n************ BIBLIOTECA X ************")
    print("             BIENVENIDO   ")
    print("este es el menu de la biblioteca X elige una de las siguientes opciones para continuar:")
    print("[1] Registra un nuevo ejemplar :")
    print("[2] ver las consultas y los reportes : ")
    print("[3] Registra un nuevo autor :")
    print("[4] Registra un nuevo genero :")
    print("[X] salir de el menu de la biblioteca : ")
    
    eleccion=input("indique su eleccion:  ").upper()
    if eleccion == "1" :
        try:
            with sqlite3.connect("biblioteca x.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * from autor;")
                autores = cursor.fetchall()
        except Error as e:
            print (e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        if autores :
            try:
                with sqlite3.connect("biblioteca x.db") as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT * from genero;")
                    generos = cursor.fetchall()
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            if generos :
                titulo=input("ingresa el nombre del libro:  ").upper()
                while titulo.isspace() or titulo=='':
                    print("\ntienes que ingresar el titulo del libro")
                    titulo=input("ingresa el nombre del libro:  ").upper()
                try:
                    with sqlite3.connect("biblioteca x.db") as conn:
                        cursor = conn.cursor()
                        titulo_nuevo={"titulo":titulo}
                        cursor.execute("select titulo from ejemplar where titulo= :titulo;", titulo_nuevo)
                        titulo_existe=cursor.fetchall()
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                if titulo_existe :
                    print("\nya existe un ingreso con ese titulo")
                else:
                    print("\n***************************")
                    print(f"ID\tAUTOR")
                    print("***************************")
                    for id, nombres, apellidos in autores:
                        print(f"{id:3}|{nombres:15}|{apellidos:20}")
                    print("***************************\n")
                    autor=int(input("ingresa el id del autor:  "))
                    while autor==0 or autor not in ((id) for id, nombres, apellidos in autores):
                        print("\ntienes que ingresar un autor valido del libro")
                        autor=int(input("ingresa el id del autor:  "))
                    print("\n***************************")
                    print(f"ID\GENERO")
                    print("***************************")
                    for id, nombre in generos:
                        print(f"{id:3}|{nombre:20}")
                    print("***************************\n")
                    genero=int(input("ingresa el id del genero:  "))
                    while genero==0 or genero not in ((id) for id, nombre in generos):
                        print("\ntienes que ingresar un genero valido del libro")
                        genero=int(input("ingresa el id del genero:  "))
                    año_pub=int(input("ingresa el año de publicacion del libro:  "))
                    while año_pub==0 :
                        print("\ntienes que ingresar el año de publicacion del libro")
                        año_pub=int(input("ingresa el año de publicacion del libro:  "))
                    isbn=input(f"ingresa el isbn de {titulo}:  ")
                    while isbn.isspace() or isbn=='' :
                        print("\ntienes que ingresar el isbn del libro")
                        isbn=int(input(f"ingresa el isbn de {titulo}:  "))
                    fech_adq=input("ingresa la fecha de adquisicion (dd/mm/aaaa):  ")
                    while fech_adq.isspace() or fech_adq=='':
                        print("\ntienes que ingresar la fecha de adquisicion del libro") 
                        fech_adq=input("ingresa la fecha de adquisicion (dd/mm/aaaa):  ")
                    try:
                        with sqlite3.connect("biblioteca x.db") as conn:
                            cursor = conn.cursor()
                            libro_datos = {"titulo":titulo, "autor":autor, "genero":genero, "año":año_pub, "isbn":isbn,"fecha":fech_adq}
                            cursor.execute("INSERT INTO ejemplar (titulo, autor, genero, año_pub, isbn, fech_adq) VALUES(:titulo,:autor,:genero,:año,:isbn,:fecha)", libro_datos)
                            print("\n********ingreso generado exitosamente********")
                    except Error as e:
                        print (e)
                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            else:
                print("no hay generos ingresados")
        else :
            print("no hay autores ingresados")
    elif eleccion == "2" :
        try:
            with sqlite3.connect("biblioteca x.db") as conn:
                cursor = conn.cursor()
                cursor.execute("select * from ejemplar;")
                ejemplares=cursor.fetchall()
                if ejemplares:
                    while True:
                        print("\n****CONSULTAS Y REPORTES****")
                        elecc=input("ELIGA UNA DE LAS SIG. OPCIONES\n[1] PARA CONSULTA DE TITULO\n[2] PARA REPORTES\n[X] PARA VOLVER AL MENU PRINCIPAL\n->")
                        if elecc == "1" :
                            while True:
                                elecc_1=input("COMO DESEAS BUSCAR EL LIBRO\n[1] POR TITULO\n[2] POR ISBN\n[X]VOLVER AL MENU DE CONSULTAS Y REPORTES\n").upper()
                                if elecc_1 == "1" :
                                    print("*********************")
                                    print(f"ID\tTITULO")
                                    print("*********************")
                                    for id,titulo,autor,genero,año_pub,isbn,fecha_adq in ejemplares:
                                        print(f"{id:4} | {titulo:30}")
                                    print("*********************")
                                    titulo_consulta=input("\ningresa el nombre del titulo para consultar todos los datos del titulo ingresado:  ").upper()
                                    try:
                                        with sqlite3.connect("biblioteca x.db") as conn:
                                            cursor = conn.cursor()
                                            libro_escogido = {"titulo":titulo_consulta}
                                            cursor.execute("SELECT e.id, e.titulo, a.nombres, a.apellidos, g.nombre, e.año_pub, e.isbn, e.fech_adq FROM ejemplar e JOIN autor a ON a.id = e.autor JOIN genero g ON g.id = e.genero WHERE e.titulo = :titulo;", libro_escogido)
                                            libro_encontrado = cursor.fetchall()
                                            if libro_encontrado :
                                                print("\n*********************")
                                                print(f"ID\tTITULO\t\t\tAUTOR\t\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN") 
                                                for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in libro_encontrado:
                                                    print(f"{id:3}|{titulo:20}|{a_nombres:10} {a_apellidos:10}|{genero:20}|{año_pub:20}|{isbn:20}|{fecha_adq:15}")
                                                print("*********************")
                                            else:
                                                print(f"\nno se encontró un título con el nombre {titulo_consulta}\n")                                                         
                                    except Error as e:
                                        print (e)
                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                
                                elif elecc_1 == "2" :
                                    print("*****************")
                                    print(f"ID\tISBN")
                                    print("******************")
                                    for id,titulo,autor,genero,año_pub,isbn,fecha_adq in ejemplares:
                                        print(f"{id:4} | {isbn}")
                                    print("*****************")
                                    isbn_consulta=input("\ningresa el isbn para consultar los titulos: ")
                                    try:
                                        with sqlite3.connect("biblioteca x.db") as conn:
                                            cursor = conn.cursor()
                                            libro_escogido = {"isbn":isbn_consulta}
                                            cursor.execute("SELECT e.id, e.titulo, a.nombres, a.apellidos, g.nombre, e.año_pub, e.isbn, e.fech_adq FROM ejemplar e JOIN autor a ON a.id = e.autor JOIN genero g ON g.id = e.genero WHERE e.isbn = :isbn", libro_escogido)
                                            libro_encontrado = cursor.fetchall()
                                            if libro_encontrado :
                                                print("\n*********************")
                                                print(f"ID\tTITULO\t\t\tAUTOR\t\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN") 
                                                for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in libro_encontrado:
                                                    print(f"{id:3}|{titulo:20}|{a_nombres:20}|{a_apellidos:20}|{genero:20}|{año_pub:20}|{isbn:20}|{fecha_adq:15}")
                                                print("*********************")
                                            else:
                                                print(f"\nno se encontró un título con el isbn {isbn_consulta}\n")
                                    except Error as e:
                                        print (e)
                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                                elif elecc_1 == "x".upper():
                                    break
                        if elecc == "2" :
                            while True:
                                print("**************************************")
                                print("****************REPORTES********************")
                                print("**************************************")
                                elecc_2=input("[1]catalago completo\n[2]reporte por autor\n[3]reporte por género\n[4]por año de publicación\n[5]Volver al menú de reportes\nindique la opcion deseada: ")
                                print("")
                                if elecc_2=="1":
                                    try:
                                        with sqlite3.connect("biblioteca x.db") as conn:
                                            cursor = conn.cursor()
                                            cursor.execute("SELECT e.id, e.titulo, a.nombres, a.apellidos, g.nombre, e.año_pub, e.isbn, e.fech_adq FROM ejemplar e JOIN autor a ON a.id = e.autor JOIN genero g ON g.id = e.genero")
                                            datos_completos = cursor.fetchall()
                                            print("\n*********************")
                                            print(f"ID\tTITULO\t\t\tAUTOR\t\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN") 
                                            for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in datos_completos:
                                                print(f"{id:3}|{titulo:20}|{a_nombres:10} {a_apellidos:10}|{genero:20}|{año_pub:20}|{isbn:20}|{fecha_adq:15}")
                                            print("*********************")
                                    except Error as e:
                                        print (e)
                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    while True:
                                        elecc_2_1=input("[1] exportar reporte en formato csv\n[2] exportar reporte en formato miexcel\n[3] no exportar reporte\n->")
                                        if elecc_2_1 == "1" :
                                            for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in datos_completos :
                                                diccionario_ejemplar[id] = [titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq]
                                            with open("ejemplares completos.csv","w",newline="") as archivo_completos:
                                                    grabador=csv.writer(archivo_completos)
                                                    grabador.writerow(("id", "titulo", "nombres de autor", "apellidos de autor", "genero", "año publicacion", "isbn", "fecha adquisicion"))
                                                    grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5], datos[6]) for clave, datos in diccionario_ejemplar.items()])
                                            print(f'reporte exportado, se ha guardado con el nombre "ejemplares completos.csv" en {os.getcwd()}\n')
                                        if elecc_2_1=="2":
                                            libro = openpyxl.Workbook()
                                            libro.iso_dates = True
                                            hoja = libro["Sheet"]
                                            hoja.title ="ejemplares"
                                            columna=0
                                            for encabezado in encabezados:
                                                columna=columna+1
                                                hoja.cell(row=1, column=columna).value = encabezado
                                            renglon=1
                                            for elemento in datos_completos :
                                                renglon=renglon+1
                                                columna=0
                                                for dato in elemento:
                                                    columna=columna+1
                                                    hoja.cell(row=renglon, column=columna).value = dato
                                            libro.save("ejemplares completos.xlsx")
                                            print(f'reporte exportado, se ha guardado con el nombre "ejemplares completos.xlsx" en {os.getcwd()}\n')
                                        if elecc_2_1=="3":
                                            break 
                                if elecc_2 == "2" :
                                    try:
                                        with sqlite3.connect("biblioteca x.db") as conn:
                                            cursor = conn.cursor()
                                            cursor.execute("SELECT * from autor;")
                                            autores = cursor.fetchall()
                                    except Error as e:
                                        print (e)
                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    if autores:
                                        print("***************************")
                                        print(f"ID\tAUTOR")
                                        print("***************************")
                                        for id, nombres, apellidos in autores:
                                            print(f"{id:3}|{nombres:20}|{apellidos:20}")
                                        print("***************************")
                                        autor_opcion=input(f"\nelija un id de autor para ver sus respectivos libros: ")
                                        try:
                                            with sqlite3.connect("biblioteca x.db") as conn:
                                                cursor = conn.cursor()
                                                autor_escogido = {"id" : autor_opcion}
                                                cursor.execute("SELECT e.id, e.titulo, a.nombres, a.apellidos, g.nombre, e.año_pub, e.isbn, e.fech_adq FROM ejemplar e JOIN autor a ON a.id = e.autor JOIN genero g ON g.id = e.genero WHERE e.autor = :id", autor_escogido)
                                                ejemplar_por_autor = cursor.fetchall()
                                        except Error as e:
                                            print (e)
                                        except:
                                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                        if ejemplar_por_autor :
                                            print("\n*********************")
                                            print(f"ID\tTITULO\t\t\tAUTOR\t\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN") 
                                            for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in ejemplar_por_autor:
                                                print(f"{id:3}|{titulo:20}|{a_nombres:10} {a_apellidos:10}|{genero:20}|{año_pub:20}|{isbn:20}|{fecha_adq:15}")
                                            print("*********************")
                                            while True:
                                                elecc_2_2=input("\n[1] exportar reporte en formato csv\n[2] exportar reporte en formato miexcel\n[3] no exportar reporte\n->")
                                                if elecc_2_2 == "1" :
                                                    for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in ejemplar_por_autor :
                                                        diccionario_ejemplar_autor[id] = [titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq]
                                                    with open("ejemplares por autor.csv","w",newline="") as archivo_autor:
                                                            grabador=csv.writer(archivo_autor)
                                                            grabador.writerow(("id", "titulo", "nombres de autor", "apellidos de autor", "genero", "año publicacion", "isbn", "fecha adquisicion"))
                                                            grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5], datos[6]) for clave, datos in diccionario_ejemplar_autor.items()])
                                                    print(f'reporte exportado, se ha guardado con el nombre "ejemplares por autor.csv" en {os.getcwd()}\n')
                                                if elecc_2_2=="2":
                                                    libro_autor = openpyxl.Workbook()
                                                    libro_autor.iso_dates = True
                                                    hoja_autor = libro_autor["Sheet"]
                                                    hoja_autor.title ="ejemplares"
                                                    columna=0
                                                    for encabezado in encabezados:
                                                        columna=columna+1
                                                        hoja_autor.cell(row=1, column=columna).value = encabezado
                                                    renglon=1
                                                    for elemento in ejemplar_por_autor :
                                                        renglon=renglon+1
                                                        columna=0
                                                        for dato in elemento:
                                                            columna=columna+1
                                                            hoja_autor.cell(row=renglon, column=columna).value = dato
                                                    libro_autor.save("ejemplares por autor.xlsx")
                                                    print(f'reporte exportado, se ha guardado con el nombre "ejemplares por autor.xlsx" en {os.getcwd()}\n')
                                                if elecc_2_2=="3":
                                                    break
                                        else:
                                            print(f"\nno se encontró un título con el autor {autor_opcion}\n")
                                    else:
                                        print("\nno hay autores ingresados")
                                if elecc_2 == "3" :
                                    try:
                                        with sqlite3.connect("biblioteca x.db") as conn:
                                            cursor = conn.cursor()
                                            cursor.execute("SELECT * from genero;")
                                            generos = cursor.fetchall()
                                    except Error as e:
                                        print (e)
                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    if generos:
                                        print("***************************")
                                        print(f"ID\GENERO")
                                        print("***************************")
                                        for id, nombre in generos:
                                            print(f"{id:3}|{nombre:20}")
                                        print("***************************")
                                        genero_opcion=input(f"\nelija un id de genero para ver sus respectivos libros: ")
                                        try:
                                            with sqlite3.connect("biblioteca x.db") as conn:
                                                cursor = conn.cursor()
                                                genero_escogido = {"id" : genero_opcion}
                                                cursor.execute("SELECT e.id, e.titulo, a.nombres, a.apellidos, g.nombre, e.año_pub, e.isbn, e.fech_adq FROM ejemplar e JOIN autor a ON a.id = e.autor JOIN genero g ON g.id = e.genero WHERE e.genero = :id", genero_escogido)
                                                ejemplar_por_genero = cursor.fetchall()
                                        except Error as e:
                                            print (e)
                                        except:
                                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                        if ejemplar_por_genero :
                                            print("\n*********************")
                                            print(f"ID\tTITULO\t\t\tAUTOR\t\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN") 
                                            for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in ejemplar_por_genero :
                                                print(f"{id:3}|{titulo:20}|{a_nombres:10} {a_apellidos:10}|{genero:20}|{año_pub:20}|{isbn:20}|{fecha_adq:15}")
                                            print("*********************")
                                            while True:
                                                elecc_2_3=input("\n[1] exportar reporte en formato csv\n[2] exportar reporte en formato miexcel\n[3] no exportar reporte\n->")
                                                if elecc_2_3 == "1" :
                                                    for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in ejemplar_por_genero :
                                                        diccionario_ejemplar_genero[id] = [titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq]
                                                    with open("ejemplares por genero.csv","w",newline="") as archivo_genero:
                                                            grabador=csv.writer(archivo_genero)
                                                            grabador.writerow(("id", "titulo", "nombres de autor", "apellidos de autor", "genero", "año publicacion", "isbn", "fecha adquisicion"))
                                                            grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5], datos[6]) for clave, datos in diccionario_ejemplar_genero.items()])
                                                    print(f'reporte exportado, se ha guardado con el nombre "ejemplares por genero.csv" en {os.getcwd()}\n')
                                                if elecc_2_3=="2":
                                                    libro_genero = openpyxl.Workbook()
                                                    libro_genero.iso_dates = True
                                                    hoja_genero = libro_genero["Sheet"]
                                                    hoja_genero.title ="ejemplares"
                                                    columna=0
                                                    for encabezado in encabezados:
                                                        columna=columna+1
                                                        hoja_genero.cell(row=1, column=columna).value = encabezado
                                                    renglon=1
                                                    for elemento in ejemplar_por_genero :
                                                        renglon=renglon+1
                                                        columna=0
                                                        for dato in elemento:
                                                            columna=columna+1
                                                            hoja_genero.cell(row=renglon, column=columna).value = dato
                                                    libro_genero.save("ejemplares por genero.xlsx")
                                                    print(f'reporte exportado, se ha guardado con el nombre "ejemplares por genero.xlsx" en {os.getcwd()}\n')
                                                if elecc_2_3=="3":
                                                    break
                                        else:
                                            print(f"\nno se encontraron registros con el género {genero_opcion}")
                                    else:
                                        print("\nno hay generos ingresados")
                                if elecc_2 == "4" :
                                    año_opcion=int(input(f"\nelija un año para ver los respectivos libros publicados: "))
                                    try:
                                        with sqlite3.connect("biblioteca x.db") as conn:
                                            cursor = conn.cursor()
                                            año_escogido = {"año" : año_opcion}
                                            cursor.execute("SELECT e.id, e.titulo, a.nombres, a.apellidos, g.nombre, e.año_pub, e.isbn, e.fech_adq FROM ejemplar e JOIN autor a ON a.id = e.autor JOIN genero g ON g.id = e.genero WHERE e.año_pub = :año", año_escogido)
                                            ejemplar_por_año = cursor.fetchall()
                                    except Error as e:
                                        print (e)
                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    if ejemplar_por_año :
                                        print("\n*********************")
                                        print(f"ID\tTITULO\t\t\tAUTOR\t\t\tGENERO\t\tAÑO DE PUBLICACION\tISBN\t\tFECHA DE ADQUISICIÓN") 
                                        for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in ejemplar_por_año :
                                            print(f"{id:3}|{titulo:20}|{a_nombres:10} {a_apellidos:10}|{genero:20}|{año_pub:20}|{isbn:20}|{fecha_adq:15}")
                                        print("*********************")
                                        while True:
                                            elecc_2_4=input("\n[1] exportar reporte en formato csv\n[2] exportar reporte en formato miexcel\n[3] no exportar reporte\n->")
                                            if elecc_2_4 == "1" :
                                                for id,titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq in ejemplar_por_año :
                                                    diccionario_ejemplar_año[id] = [titulo,a_nombres,a_apellidos,genero,año_pub,isbn,fecha_adq]
                                                with open("ejemplares por año.csv","w",newline="") as archivo_año:
                                                        grabador=csv.writer(archivo_año)
                                                        grabador.writerow(("id", "titulo", "nombres de autor", "apellidos de autor", "genero", "año publicacion", "isbn", "fecha adquisicion"))
                                                        grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5], datos[6]) for clave, datos in diccionario_ejemplar_año.items()])
                                                print(f'reporte exportado, se ha guardado con el nombre "ejemplares por año.csv" en {os.getcwd()}\n')
                                            if elecc_2_4=="2":
                                                libro_año = openpyxl.Workbook()
                                                libro_año.iso_dates = True
                                                hoja_año = libro_año["Sheet"]
                                                hoja_año.title ="ejemplares"
                                                columna=0
                                                for encabezado in encabezados:
                                                    columna=columna+1
                                                    hoja_año.cell(row=1, column=columna).value = encabezado
                                                renglon=1
                                                for elemento in ejemplar_por_año :
                                                    renglon=renglon+1
                                                    columna=0
                                                    for dato in elemento:
                                                        columna=columna+1
                                                        hoja_año.cell(row=renglon, column=columna).value = dato
                                                libro_año.save("ejemplares por año.xlsx")
                                                print(f'reporte exportado, se ha guardado con el nombre "ejemplares por año.xlsx" en {os.getcwd()}\n')
                                            if elecc_2_4=="3":
                                                break
                                if elecc_2 == "5" :
                                    break
                        if elecc == "x" :
                            break
                else:
                    print("\nno hay títulos registrados por el momento") 
        except Error as e:
            print (e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

    elif eleccion== "3" :
        autor_nombres=input("ingresa el nombre del autor:  ").upper()
        while autor_nombres.isspace() or autor_nombres=='' :
            print("\ntienes que ingresar los nombres del autor\n")
            autor_nombres=input("ingresa el nombre del autor:  ").upper()
        autor_apellidos=input("ingresa los apellidos del autor:  ").upper()
        while autor_apellidos.isspace() or autor_apellidos=='':
            print("\ntienes que ingresar los apellidos del autor\n")
            autor_apellidos=input("ingresa los apellidos del autor:  ").upper()
        try:
            with sqlite3.connect("biblioteca x.db") as conn:
                cursor = conn.cursor()
                autor_datos = {"nombres":autor_nombres, "apellidos":autor_apellidos}
                cursor.execute("select * from autor WHERE nombres = :nombres AND apellidos = :apellidos", autor_datos)
                autor_ingresado=cursor.fetchall()
                if autor_ingresado:
                    print("\nel autor ya está registrado!\n")
                else:
                    cursor.execute("INSERT INTO autor (nombres, apellidos) VALUES(:nombres,:apellidos)", autor_datos)
                    print("\n********ingreso generado exitosamente********")
        except Error as e:
            print (e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    elif eleccion == "4" :
        genero_nombre=input("ingresa el nombre del genero:  ").upper()
        while genero_nombre.isspace() or genero_nombre=='':
            print("\ntienes que ingresar un nombre de genero\n")
            genero_nombre=input("ingresa el nombre del genero:  ").upper()
        try:
            with sqlite3.connect("biblioteca x.db") as conn:
                cursor = conn.cursor()
                genero_datos = {"nombre":genero_nombre}
                cursor.execute("select * from genero WHERE nombre = :nombre", genero_datos)
                genero_ingresado=cursor.fetchall()
                if genero_ingresado:
                    print("\nel género ya está registrado!\n")
                else:
                    cursor.execute("INSERT INTO genero (nombre) VALUES(:nombre)", genero_datos)
                    print("\n********ingreso generado exitosamente********")
        except Error as e:
            print (e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

    elif eleccion == "x".upper() :
        print("SYSTEM OFF")
        break